"""
EXCEL FILES:
-Openpyxl allows your Python programs to read, edit and create Excel spreadsheet files.

PDF AND WORD FILES:
-PyPDF2 allows your Python programs to read, edit and create PDF files.
-Docx allows your Python programs to read, edit and create Word files.

CSV FILES:
-CSV stands for “comma-separated values,” and CSV files are simplified spreadsheets stored as plaintext files.
-CSV module makes it easy to parse CSV files.

JSON FILES:
-JSON (JavaScript Object Notation) is a popular way to format data as a single human-readable string.
-JSON is the native way that JavaScript programs write their data structures.
-Many websites offer JSON content as a way for programs to interact with the website known as application programming
interface (API).
"""

import openpyxl
import csv
import json
from openpyxl.utils import get_column_letter, column_index_from_string



# -----------------EXCEL FILES-----------------

# Read excel files
workbook = openpyxl.load_workbook('Files\\example.xlsx')

# Workbook attributes
print(workbook.sheetnames)
print(workbook.active)

# Sheet attributes
# sheet.title       -> return title of an excel sheet
# sheet.max_row     -> return the highest row number
# sheet.max_column  -> return the highest column number
sheet = workbook['Sheet3']
print(sheet.title)
print(sheet.max_row, sheet.max_column)

# Access cells from a sheet and get their properties
# cell.value      -> return value of a cell
# cell.coordinate -> return coordinates of a cell
# cell.row        -> return row number of a cell
# cell.column     -> return column number of a cell
cellA1 = sheet['A1']
print(cellA1.value, sheet['B1'].value)
print('Row %s, Column %s is %s' % (cellA1.row, cellA1.column, cellA1.value))
print('Cell %s is %s' % (cellA1.coordinate, cellA1.value))

# Alternative way to access cells
# cell = sheet_obj.cell(row=n, column=m)
cell3 = sheet.cell(row=1, column=2)
for i in range(1, 8, 2):
    print(i, sheet.cell3(row=i, column=2).value)

# Convert between column letters to numbers
column_letter = get_column_letter(1)          # Translate column 1 to a letter
column_index = column_index_from_string('A')  # Get A's number

# Get rows and columns from the Sheets
for row in sheet['A1':'C3']:
    for cell in row:
        print(cell.coordinate, cell.value)
    print('--- END OF ROW ---')






# -----------------CSV FILES-----------------

# Read csv files
exampleFile1 = open('.\\Files\\example.csv')
exampleReader1 = csv.reader(exampleFile1)

# Access data in a csv file
exampleData = list(exampleReader1)
print(exampleData[0][0], exampleData[4][1])

# Close csv file after using it
exampleFile1.close()

# For larger csv files, it is better to loop over them
# line_num -> return the number of a line
exampleFile2 = open('.\\Files\\example.csv')
exampleReader2 = csv.reader(exampleFile2)
for row in exampleReader2:
    print('Row #' + str(exampleReader2.line_num) + ' ' + str(row))
exampleFile2.close()

# Write csv files
# csv.writer writes \r\n into the file directly
# In Python 2, open outfile with mode 'wb' instead of 'w' to avoid blank lines
# In Python 3, open outfile with the additional parameter newline=''
# newline        -> parameter to set newline separator, \n by default
# delimiter      -> parameter to set delimiter, \t by default
# lineterminator -> parameter to set separators between lines
outputFile = open('.\\Files\\output.csv', 'w', newline='')
outputWriter = csv.writer(outputFile, delimiter='\t', lineterminator='\n\n')

# Write rows to a csv writer object
outputWriter.writerow(['spam', 'eggs', 'bacon', 'ham'])
outputFile.close()






# -----------------CSV FILES - DICTREADER & DICTWRITER-----------------

# They use a dictionaries instead of lists to process csv files
# They are more convenient for csv files with header rows

# Read csv files with DictReader
exampleFile3 = open('.\\Files\\exampleWithHeader.csv')
exampleDictReader3 = csv.DictReader(exampleFile3)
for row in exampleDictReader3:
    print(row['Timestamp'], row['Fruit'], row['Quantity'])
exampleFile3.close()

# Add header names for csv files without them
exampleFile4 = open('example.csv')
exampleDictReader4 = csv.DictReader(exampleFile4, ['time', 'name', 'amount'])
for row in exampleDictReader4:
    print(row['time'], row['name'], row['amount'])
exampleFile4.close()

# Create csv files with DictWriter
outputFile2 = open('output.csv', 'w', newline='')
outputDictWriter = csv.DictWriter(outputFile2, ['Name', 'Pet', 'Phone'])
# Write header to csv file
outputDictWriter.writeheader()
# Write rows to csv file
outputDictWriter.writerow({'Name': 'Alice', 'Pet': 'cat', 'Phone': '555-1234'})
outputDictWriter.writerow({'Name': 'Bob', 'Phone': '555-9999'})
outputFile2.close()






# -----------------JSON FILES-----------------

# load() -> read data from json file and return it as a dictionary object
JsonData_string = '{"name": "Zophie", "isCat": true, "miceCaught": 0, "felineIQ": null}'
JsonData_python = json.loads(JsonData_string)

# dump() -> write data to json file
pythonValue = {'isCat': True, 'miceCaught': 0, 'name': 'Zophie', 'felineIQ': None}
stringJsonData = json.dumps(pythonValue)
