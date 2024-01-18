"""
EXCEL FILES:
-Openpyxl allows your Python programs to read, edit and create Excel spreadsheet files.

PDF AND WORD FILES:
-PyPDF2 allows your Python programs to read, edit and create PDF files.
-Docx allows your Python programs to read, edit and create Word files.

CSV FILES:
-CSV stands for “comma-separated values,” and CSV files are simplified spreadsheets stored as
plaintext files.
-CSV module makes it easy to parse CSV files.

JSON FILES:
-JSON (JavaScript Object Notation) is a popular way to format data as a single human-readable
string.
-JSON is the native way that JavaScript programs write their data structures.
-Many websites offer JSON content as a way for programs to interact with the website known as
application programming interface (API).
"""

import csv
import json
import openpyxl
import os
import shutil
import zipfile as zf

from openpyxl.utils import get_column_letter, column_index_from_string
from PIL import Image
from pathlib import Path


# -----------------FILE MANAGEMENT-----------------

# Create files using open function
# open(filename, 'file_mode') -> open a file in a given file mode
# add r before file path avoids special characters (like \n). Alternatively, we can use \\
filename = (
    r'C:\Users\ponce\Documents\Education\Degree - Software Engineering\Python\Course - Python '
    r'3\12-Advanced Python Modules\practice.txt '
)
f = open(filename, 'w+')

# Create files using with statement
# Using "with", the file will automatically be closed
with open(filename, 'w+') as my_file:
    my_file.close()

# File modes
# 'w'  -> writing mode
# 'w+' -> writing and reading mode, overwrites the existing file
# 'wb' -> write and binary permissions
# 'r'  -> reading mode
# 'r+' -> reading and writing mode
# 'a'  -> appending mode

# File operations
# read()          -> return the content of a file
# readlines()     -> return a list of strvalues from the file, one string for each line of text
# write('string') -> write some string on a file
# close()         -> close a file
print(f.read())
print(f.readlines())
f.write('test')
f.close()

# Remove files and directories
# os.unlink/remove(path) -> delete a file at the path your provide. It can't be reversed
# os.rmdir(path)         -> delete an empty dir at the path your provide. It can't be reversed
# shutil.rmtree(path)    -> delete all files and folders in the path. It can't be reversed
# send2trash(path)       -> send files to trash can instead of remove them
os.remove(filename)

# Grant permissions
# os.chmod(path, mode) -> change permissions of path using given mode that can be an octal
# (for example, Oo644) or a stat variable from stat module
os.chmod(filename, 0o644)




# -----------------PATHS AND DIRECTORIES-----------------

# os.getcwd() -> return current working directory
# os.chdir()  -> change current working directory
print(os.getcwd())
os.chdir('C:\\Windows\\System32')

# os.makedirs() -> create a folder in the given directory
os.makedirs('C:\\Users\\ponce\\Documents\\test')

# os.listdir() -> list files in the current or a given directory as a list of strings
print(os.listdir())
print(os.listdir(r"C:\Users"))

# Check files or directories
# os.path.isabs(path)  -> return True if the argument is an absolute path
# os.path.isfile(path) -> return True if the argument is an existing regular file
# os.path.isdir(path)  -> return True if the argument is an existing directory
print(os.path.isabs(os.getcwd()))
print(os.path.isfile(filename))
print(os.path.isdir(filename))

# Get paths as strings
# os.path.abspath(path)        -> return a str of the absolute path of the argument
# os.path.relpath(path, start) -> return a str of a relative path from the start
# os.path.basename(path)       -> return a str of everything that comes after last \ in a path
# os.path.dirname(path)        -> return a str of everything that comes before last \ in a path
# os.path.split()              -> return a tuple with path’s dir name and base name as str
os.path.abspath('.\\Games')  # .  -> shorthand for the current directory
os.path.isabs('..\\Games')   # .. -> shorthand for the parent folder
os.path.relpath('C:\\Windows', 'C:\\')

# Get status of a file or directory
# os.stat(path) -> performs stat() system call on the specified path, so we get its status
# The returned ‘stat-result’ object has the following attributes:
# st_mode  -> represents file type and file mode bits (permissions)
# st_dev   -> represents the identifier of the device on which this file resides
# st_uid   -> represents the user identifier of the file owner
# st_gid   -> represents the group identifier of the file owner
# st_size  -> represents the size of the file in bytes
# st_atime -> represents the time of most recent access in seconds
# st_mtime -> represents the time of most recent content modification in seconds
# st_ctime -> represents the time of most recent metadata change on Unix and creation time on
# Windows in seconds
status = os.stat(os.getcwd())

# Move and copy directories with shutil library
# shutil.move(filename, path)   -> move or rename files
# shutil.copy(filename, path)   -> copy files or folders
# shutil.copytree(folder, path) -> copy an entire folder (including files and subfolders)
shutil.move(filename, r'C:\Users\ponce')
shutil.copy(filename, r'C:\Users\ponce')
shutil.copy(filename, r'C:\Users\ponce\new_name.txt')
shutil.copytree(r'C:\Users\ponce\Documents\Personal', r'C:\Users\ponce')

# Walk through a directory
# os.walk(path) -> generate the file names in a directory tree by walking the tree either
# top-down or bottom-up
for root, sub_folders, files in os.walk(r'C:\Users\ponce\Downloads'):
    print("Currently looking at folder: " + root)
    print('Subfolders: ')
    for sub_fold in sub_folders:
        print('\t Subfolder: ' + sub_fold)
    print('Files are: ')
    for f in files:
        print('\t File: ' + f)
    print('\n')




# ------------------PATH LIBRARY (ONLY PYTHON 3.5+)-----------------

# Path.cwd()   -> return current working directory
# Path.home()  -> return home directory
# Path().mkdir ->  create a folder in the given directory
print(Path.cwd())
print(Path.home())
Path('C:\\Users\\ponce\\Documents\\test').mkdir()

# Getting the Parts of a File Path
# anchor  -> root folder of the filesystem
# drive   -> single letter that often denotes a physical storage
# name    -> a string representing the final path component, excluding the drive and root
# stem    -> final path component, without its suffix
# suffix  -> file extension of the final component
# parent  -> folder that contains the file
# parents -> evaluates to the ancestor folders of a Path object with an integer index
# The name of the file, made up of the stem (or base name) and the suffix (or extension)
p = Path('C:\\Users\\ponce\\Documents')

# Navigating inside a directory tree:
file = p / "example.txt"

print(file.anchor, file.name, file.stem, file.suffix, file.drive)
print(Path.cwd().parents[0])
print(Path.cwd().parents[2])

# glob() -> list the contents of a folder according to a pattern
# glob returns a generator object, that can be turned into a list
p = Path('C:\\Users\\ponce\\Desktop')
print(list(p.glob('*')))  # the asterisk stands for "multiple of any characters"
print(list(p.glob('*.txt')))  # lists all text files

# exists()  -> return True if the path exists
# is_file() -> return True if the path exists and is a file
# is_dir()  -> return True if the path exists and is a directory
winDir = Path('C:\\Windows')
calcFile = Path('C:\\Windows\\System32\\calc.exe')
winDir.exists()
calcFile.is_file()
calcFile.is_dir()




# ------------------ZIPPING FILES-----------------

# zipfile() -> zip files to compress them
# To compress a file, create a zip file object, then write to it (the write step compresses
# the files)
# To compress all files in a folder, just use the os.walk() method to iterate this process
compressed_file = zf.ZipFile('comp_file.zip', 'w')
compressed_file.write("new_file.txt", compress_type=zf.ZIP_DEFLATED)
compressed_file.close()

# extractall() -> extract files in a folder
# extract()    -> extract individual file
zip_obj = zf.ZipFile('comp_file.zip', 'r')
zip_obj.extractall("extracted_content")

# shutil.make_archive()   -> archive all files in a folder
# shutil.unpack_archive() -> extract all files in a folder
directory_to_zip = r'C:\Users\ponce\Documents\Personal\Healthcare'
output_filename = 'example'
shutil.make_archive(output_filename, 'zip', directory_to_zip)




# -----------------WORKING WITH IMAGES-----------------

# Open/save images
mac = Image.open('example.jpg')
mac.save("laptop.png")

# Image information
print(type(mac))
print(mac.size)
print(mac.info)
print(mac.format_description)

# copy() -> copy the image to another image object
computer = mac.copy()

# paste(image1, image2, box, mask) -> paste an image on another image
# image1 is the image on which other image is to be pasted (optional)
# image2 is the original image
# box is a 4-tuple giving the region to paste into (optional)
# mask is an optional mask image
mac.paste(im=computer, box=(796, 0))

# crop((x,y,w,h)) -> return a rectangular region from an image
# x, y indicate the initial coordinate of the crop rectangle. They start at top corner (0,0)
# w, h indicate the width and height of the crop rectangle
# Start at top corner (0,0)
x = 0
y = 0
w = int(1950/3)
h = int(1300/10)
print(mac.crop((x, y, w, h)))

# resize((new_h,new_w)) -> method to resize an image
mac.resize((1000, 500))

# rotate() -> rotate a given image to the given number of degrees counterclockwise
# expand is optional and expands the output image to make it large enough to hold the entire
# rotated image
mac.rotate(90, expand=True)




# -----------------EXCEL FILES-----------------

# Read excel files
workbook = openpyxl.load_workbook('Files\\example.xlsx')

# Workbook attributes
print(workbook.sheetnames)
print(workbook.active)

# Sheet attributes
# sheet.title       -> return title of an Excel sheet
# sheet.max_row     -> return the highest row number
# sheet.max_column  -> return the highest column number
sheet = workbook['Sheet3']
print(sheet.title)
print(sheet.max_row, sheet.max_column)

# Access cells from a sheet and get their properties
# cell.value      -> return value of a cell
# cell.coordinate -> return coordinates of a cell
# cell.row        -> return row number of a cell
# cell.column     -> return column number of a cell
cellA1 = sheet['A1']
print(cellA1.value, sheet['B1'].value)
print('Row %s, Column %s is %s' % (cellA1.row, cellA1.column, cellA1.value))
print('Cell %s is %s' % (cellA1.coordinate, cellA1.value))

# Alternative way to access cells
# cell = sheet_obj.cell(row=n, column=m)
cell3 = sheet.cell(row=1, column=2)
for i in range(1, 8, 2):
    print(i, sheet.cell3(row=i, column=2).value)

# Convert between column letters to numbers
column_letter = get_column_letter(1)          # Translate column 1 to a letter
column_index = column_index_from_string('A')  # Get A's number

# Get rows and columns from the Sheets
for row in sheet['A1':'C3']:
    for cell in row:
        print(cell.coordinate, cell.value)
    print('--- END OF ROW ---')




# -----------------CSV FILES-----------------

# Read csv files
exampleFile1 = open('.\\Files\\example.csv')
exampleReader1 = csv.reader(exampleFile1)

# Access data in a csv file
exampleData = list(exampleReader1)
print(exampleData[0][0], exampleData[4][1])

# Close csv file after using it
exampleFile1.close()

# For larger csv files, it is better to loop over them
# line_num -> return the number of a line
exampleFile2 = open('.\\Files\\example.csv')
exampleReader2 = csv.reader(exampleFile2)
for row in exampleReader2:
    print('Row #' + str(exampleReader2.line_num) + ' ' + str(row))
exampleFile2.close()

# Write csv files
# csv.writer writes \r\n into the file directly
# In Python 2, open outfile with mode 'wb' instead of 'w' to avoid blank lines
# In Python 3, open outfile with the additional parameter newline=''
# newline        -> parameter to set newline separator, \n by default
# delimiter      -> parameter to set delimiter, \t by default
# lineterminator -> parameter to set separators between lines
outputFile = open('.\\Files\\output.csv', 'w', newline='')
outputWriter = csv.writer(outputFile, delimiter='\t', lineterminator='\n\n')

# Write rows to a csv writer object
outputWriter.writerow(['spam', 'eggs', 'bacon', 'ham'])
outputFile.close()




# -----------------CSV FILES - DICTREADER & DICTWRITER-----------------

# They use a dictionaries instead of lists to process csv files
# They are more convenient for csv files with header rows

# Read csv files with DictReader
exampleFile3 = open('.\\Files\\exampleWithHeader.csv')
exampleDictReader3 = csv.DictReader(exampleFile3)
for row in exampleDictReader3:
    print(row['Timestamp'], row['Fruit'], row['Quantity'])
exampleFile3.close()

# Add header names for csv files without them
exampleFile4 = open('example.csv')
exampleDictReader4 = csv.DictReader(exampleFile4, ['time', 'name', 'amount'])
for row in exampleDictReader4:
    print(row['time'], row['name'], row['amount'])
exampleFile4.close()

# Create csv files with DictWriter
outputFile2 = open('output.csv', 'w', newline='')
outputDictWriter = csv.DictWriter(outputFile2, ['Name', 'Pet', 'Phone'])
# Write header to csv file
outputDictWriter.writeheader()
# Write rows to csv file
outputDictWriter.writerow({'Name': 'Alice', 'Pet': 'cat', 'Phone': '555-1234'})
outputDictWriter.writerow({'Name': 'Bob', 'Phone': '555-9999'})
outputFile2.close()




# -----------------JSON FILES-----------------

# load() -> read data from json file and return it as a dictionary object
JsonData_string = '{"name": "Zophie", "isCat": true, "miceCaught": 0, "felineIQ": null}'
JsonData_python = json.loads(JsonData_string)

# dump() -> write data to json file
pythonValue = {'isCat': True, 'miceCaught': 0, 'name': 'Zophie', 'felineIQ': None}
stringJsonData = json.dumps(pythonValue)
